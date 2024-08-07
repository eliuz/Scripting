from datetime import date
import os
from openpyxl import load_workbook
from openpyxl import Workbook
from tkinter import *
from tkinter import filedialog as fd
from tkinter import messagebox
import xml.etree.ElementTree as ET
import hashlib
from xml.dom import minidom

# from xlrd.formula import sheetrange
# CONVERSION DE METADATA A XML DE DIFERENTES CLIENTES SEGUN NOMBRE DE ARCHIVO,INCLUYE HASH PARA VIDEOS E IMÁGENES


class Metadata:

    def __init__(self):

        global window
        window = Tk()
        window.geometry("400x100")
        btn = Button(window, text="Start", width=42, height=5, command=self.start)
        btn.grid()
        btn.place(relx=0.5, rely=0.5, anchor=CENTER)
        self.start()
        window.mainloop()

    def start(self):

        try:
            filename = fd.askopenfile(
                title="Abrir excel de metadata", filetypes=(("Excel XLSX", "*.xlsx"),)
            )
            name = filename.name

            if "MC" in name:
                self.multicinema(filename)
            elif "MBA" in name:
                key = "MBA"
                self.multipremier(filename, key)

            elif "MP-TIGO" in name:
                key = "TIGO"
                self.multipremier(filename, key)

            elif "MP-CABLEONDA" in name:
                key = "CABLEONDA"
                self.multipremier(filename, key)

            elif "MP-MEGACABLE" in name:
                key = "MEGACABLE"
                self.multipremier(filename, key)

            elif "CL-MEGACABLE" in name:
                key = "MEGACABLE"
                print("entro")
                self.cineLatino(filename, key)

            elif "CL-CABLEONDA" in name:
                key = "CABLEONDA"
                self.cineLatino(filename, key)

            elif "CL-TIGO" in name:
                key = "TIGO"
                self.cineLatino(filename, key)

            elif "CL-DISH" in name:
                key = "DISH"
                self.cineLatino(filename, key)

            elif "CL-YOUTUBE" in name:
                key = "YOUTUBE"
                self.cineLatino(filename, key)

            else:
                messagebox.showwarning(
                    "", "Coloque el canal y el cliente en el nombre del archivo"
                )

        except Exception as e:
            messagebox.showerror("error Seleccione correctamente el archivo", e)

    def multicinema(self, filename):

        try:

            messagebox.showinfo("", "Abrir donde se guardarán los datos exportados ")
            # DIRECTORIO para exportados
            Exporta = fd.askdirectory(
                title="Abrir donde se guardarán los datos exportados de Multi Cinema"
            )
            path1 = os.path.join(Exporta, "MEGACABLE")
            os.mkdir(path1)
            nameProveedor = "MC"
            # Lectura de xlsx
            book = load_workbook(filename=filename.name)
            sheet = book.active
            # contador 1 para ID
            contador = 11

            for fila in range(3, sheet.max_row + 1):

                houseNumber = (sheet.cell(row=fila, column=1)).value
                ID = houseNumber[2:] + "000000000000" + str(contador)
                title = (sheet.cell(row=fila, column=2)).value
                original = (sheet.cell(row=fila, column=3)).value
                sinopsis = (sheet.cell(row=fila, column=4)).value
                actor1 = (sheet.cell(row=fila, column=5)).value
                actor2 = (sheet.cell(row=fila, column=6)).value
                actor3 = (sheet.cell(row=fila, column=7)).value
                director = (sheet.cell(row=fila, column=8)).value
                genero = (sheet.cell(row=fila, column=9)).value
                country = (sheet.cell(row=fila, column=10)).value
                year = int((sheet.cell(row=fila, column=11)).value)
                clasif = str((sheet.cell(row=fila, column=12)).value)
                durac = str((sheet.cell(row=fila, column=13)).value)
                durac = str(durac)
                duracionDisplay = str(durac)[:-3]
                formato = sheet.cell(row=fila, column=14).value
                if "HD" in formato:
                    formato = "Y"
                else:
                    formato = "N"
                iVOD = sheet.cell(row=fila, column=15).value
                iVOD = str(iVOD.date())
                fVOD = sheet.cell(row=fila, column=16).value
                fVOD = str(fVOD.date())

                self.megacable_meta(
                    ID,
                    nameProveedor,
                    path1,
                    houseNumber,
                    country,
                    title,
                    sinopsis,
                    actor1,
                    actor2,
                    actor3,
                    director,
                    genero,
                    year,
                    clasif,
                    durac,
                    duracionDisplay,
                    formato,
                    iVOD,
                    fVOD,
                )

                contador = contador + 1
                if contador == 99:
                    contador = 11

            window.title("Parseo de Metadata")
            messagebox.showinfo("OK", "Metadata Multicinema completada")

        except Exception as e:
            messagebox.showerror("error en archivo Multicinema", e)

    def cineLatino(self, filename, key):

        try:
            messagebox.showinfo("", "selecciona carpeta con imagenes y media ")
            self.getImg = fd.askdirectory(
                title="selecciona carpeta con imagenes y media"
            )

            messagebox.showinfo("", "Carpeta donde se guardarán los datos exportados ")
            Exporta = fd.askdirectory(
                title="Carpeta para datos exportados de Cinelatino - " + key
            )
            path1 = os.path.join(Exporta, key)
            os.mkdir(path1)

            nameProveedor = "Cine Latino"
            webProveedor = "www.cinelatino.com"
            callProveedor = "CL"
            openFile = load_workbook(filename=filename.name)
            sheet = openFile.active
            contador = 21

            for fila in range(3, sheet.max_row + 1):

                houseNumber = (sheet.cell(row=fila, column=1)).value

                if houseNumber.endswith("X") and len(houseNumber) == 7:
                    houseNumber = houseNumber[:6]

                print(houseNumber)
                ID = houseNumber[2:] + "000000000000" + str(contador)
                title = (sheet.cell(row=fila, column=2)).value
                sinopsis = (sheet.cell(row=fila, column=3)).value
                clasif = str((sheet.cell(row=fila, column=4)).value)
                durac = (sheet.cell(row=fila, column=5)).value
                durac = str(durac)
                duracionDisplay = str(durac)[:-3]
                year = (sheet.cell(row=fila, column=6)).value
                country = (sheet.cell(row=fila, column=7)).value
                actor1 = (sheet.cell(row=fila, column=8)).value
                actor2 = (sheet.cell(row=fila, column=9)).value
                actor3 = (sheet.cell(row=fila, column=10)).value
                director = (sheet.cell(row=fila, column=11)).value
                genero = (sheet.cell(row=fila, column=12)).value
                iVOD = (sheet.cell(row=fila, column=13)).value
                i2VOD = str(iVOD)
                iVOD = str(iVOD.date())
                fVOD = (sheet.cell(row=fila, column=14)).value
                f2VOD = str(fVOD)
                fVOD = str(fVOD.date())
                vM = str((sheet.cell(row=fila, column=15)).value)
                cc = str((sheet.cell(row=fila, column=16)).value)
                formato = "Y"

                if key == "TIGO":
                    self.tigo_meta(
                        ID,
                        self.getImg,
                        callProveedor,
                        country,
                        nameProveedor,
                        webProveedor,
                        path1,
                        houseNumber,
                        title,
                        sinopsis,
                        actor1,
                        actor2,
                        actor3,
                        director,
                        genero,
                        year,
                        clasif,
                        durac,
                        duracionDisplay,
                        formato,
                        iVOD,
                        fVOD,
                    )
                elif key == "MEGACABLE":
                    self.megacable_meta(
                        ID,
                        nameProveedor,
                        path1,
                        houseNumber,
                        country,
                        title,
                        sinopsis,
                        actor1,
                        actor2,
                        actor3,
                        director,
                        genero,
                        year,
                        clasif,
                        durac,
                        duracionDisplay,
                        formato,
                        iVOD,
                        fVOD,
                    )
                elif key == "CABLEONDA":
                    self.cableonda_meta(
                        ID,
                        self.getImg,
                        callProveedor,
                        nameProveedor,
                        webProveedor,
                        path1,
                        houseNumber,
                        title,
                        sinopsis,
                        actor1,
                        actor2,
                        actor3,
                        director,
                        genero,
                        year,
                        clasif,
                        durac,
                        duracionDisplay,
                        formato,
                        i2VOD,
                        f2VOD,
                    )
                elif key == "DISH":
                    self.dish_meta(
                        ID,
                        self.getImg,
                        callProveedor,
                        nameProveedor,
                        webProveedor,
                        path1,
                        houseNumber,
                        title,
                        sinopsis,
                        actor1,
                        actor2,
                        actor3,
                        director,
                        genero,
                        year,
                        clasif,
                        durac,
                        duracionDisplay,
                        formato,
                        iVOD,
                        fVOD,
                    )
                elif key == "YOUTUBE":
                    self.youtube_meta(
                        ID,
                        self.getImg,
                        callProveedor,
                        nameProveedor,
                        "cinelatino.com",
                        path1,
                        houseNumber,
                        title,
                        sinopsis,
                        actor1,
                        actor2,
                        actor3,
                        director,
                        genero,
                        year,
                        clasif,
                        durac,
                        duracionDisplay,
                        formato,
                        iVOD,
                        fVOD,
                        vM,
                        cc,
                    )

                else:
                    messagebox.showwarning("WARNING", "Metadata NO IDENTIFICADA")

                contador = contador + 1
                if contador == 99:
                    contador = 21

            window.title("Parseo de Metadata")
            messagebox.showinfo("OK", "Metadata Cinelatino completada")

        except Exception as e:
            messagebox.showerror("error en archivo Cinelatino", e)

    def multipremier(self, filename, key):

        try:

            messagebox.showinfo("", "selecciona carpeta con imagenes y media")
            self.getImg = fd.askdirectory(
                title="selecciona carpeta con imagenes y media"
            )

            messagebox.showinfo("", "Carpeta donde se guardarán los datos exportados ")
            Exporta = fd.askdirectory(
                title="Carpeta para datos exportados de Multipremier - " + key
            )
            path1 = os.path.join(Exporta, key)
            os.mkdir(path1)

            nameProveedor = "Multipremier"
            webProveedor = "www.multipremier.com"
            callProveedor = "MP"

            contador = 31

            openFile = load_workbook(filename=filename.name)
            sheet = openFile.active

            for fila in range(3, sheet.max_row + 1):

                houseNumber = (sheet.cell(row=fila, column=1)).value
                if houseNumber.endswith("X") and len(houseNumber) == 7:
                    houseNumber = houseNumber[:6]

                ID = houseNumber[2:] + "000000000000" + str(contador)

                title = (sheet.cell(row=fila, column=2)).value
                original = (sheet.cell(row=fila, column=3)).value
                sinopsis = (sheet.cell(row=fila, column=4)).value
                actor1 = (sheet.cell(row=fila, column=5)).value
                actor2 = (sheet.cell(row=fila, column=6)).value
                actor3 = (sheet.cell(row=fila, column=7)).value
                director = (sheet.cell(row=fila, column=8)).value
                genero = (sheet.cell(row=fila, column=9)).value
                country = (sheet.cell(row=fila, column=10)).value
                year = int((sheet.cell(row=fila, column=11)).value)
                clasif = str((sheet.cell(row=fila, column=12)).value)
                durac = str((sheet.cell(row=fila, column=13)).value)
                duracionDisplay = str(durac)[:-3]
                formato = (sheet.cell(row=fila, column=14)).value

                if "HD" in formato:
                    formato = "Y"
                else:
                    formato = "N"

                iVOD = (sheet.cell(row=fila, column=15)).value
                i2VOD = str(iVOD)
                iVOD = str(iVOD.date())
                fVOD = (sheet.cell(row=fila, column=16)).value
                f2VOD = str(fVOD)
                fVOD = str(fVOD.date())

                if key == "CABLEONDA":
                    self.cableonda_meta(
                        ID,
                        self.getImg,
                        callProveedor,
                        nameProveedor,
                        webProveedor,
                        path1,
                        houseNumber,
                        title,
                        sinopsis,
                        actor1,
                        actor2,
                        actor3,
                        director,
                        genero,
                        year,
                        clasif,
                        durac,
                        duracionDisplay,
                        formato,
                        i2VOD,
                        f2VOD,
                    )

                elif key == "MEGACABLE":
                    self.megacable_meta(
                        ID,
                        nameProveedor,
                        path1,
                        houseNumber,
                        country,
                        title,
                        sinopsis,
                        actor1,
                        actor2,
                        actor3,
                        director,
                        genero,
                        year,
                        clasif,
                        durac,
                        duracionDisplay,
                        formato,
                        iVOD,
                        fVOD,
                    )

                elif key == "TIGO":
                    self.tigo_meta(
                        ID,
                        self.getImg,
                        callProveedor,
                        country,
                        nameProveedor,
                        webProveedor,
                        path1,
                        houseNumber,
                        title,
                        sinopsis,
                        actor1,
                        actor2,
                        actor3,
                        director,
                        genero,
                        year,
                        clasif,
                        durac,
                        duracionDisplay,
                        formato,
                        iVOD,
                        fVOD,
                    )

                elif key == "MBA":
                    self.mba_meta(
                        ID,
                        self.getImg,
                        callProveedor,
                        nameProveedor,
                        webProveedor,
                        path1,
                        houseNumber,
                        title,
                        sinopsis,
                        actor1,
                        actor2,
                        actor3,
                        director,
                        genero,
                        year,
                        clasif,
                        durac,
                        duracionDisplay,
                        formato,
                        country,
                        i2VOD,
                        f2VOD,
                    )

                contador = contador + 1
                if contador == 99:
                    contador = 31

            window.title("Parseo de Metadata")
            messagebox.showinfo("OK", "Metadata Multipremier completada")

        except Exception as e:
            messagebox.showerror("error en archivo Multipremier", e)

    def mba_meta(
        self,
        ID,
        getImg,
        callProveedor,
        nameProveedor,
        webProveedor,
        path3,
        houseNumber,
        title,
        sinopsis,
        actor1,
        actor2,
        actor3,
        director,
        genero,
        year,
        clasif,
        durac,
        duracionDisplay,
        formato,
        country,
        i2VOD,
        f2VOD,
    ):

        today = date.today()
        hoy = today.strftime("%Y-%m-%d")
        if clasif == "C" or clasif == "D":
            adult = "Y"
        else:
            adult = "N"

        i = 10

        root2 = ET.Element("ADI")
        tag1 = ET.SubElement(root2, "Metadata")
        ET.SubElement(
            tag1,
            "AMS",
            Version_Minor="0",
            Version_Major="1",
            Provider_ID=webProveedor,
            Provider=nameProveedor,
            Product="MOD",
            Description=title + " package asset",
            Creation_Date=hoy,
            Asset_Name=title + " package",
            Asset_ID=ID + str(i),
            Asset_Class="package",
        )
        ET.SubElement(
            tag1,
            "App_Data",
            Value="CableLabsVOD1.1",
            Name="Metadata_Spec_Version",
            App="MOD",
        )
        ET.SubElement(
            tag1,
            "App_Data",
            Value=callProveedor + "1",
            Name="Provider_Content_Tier",
            App="MOD",
        )
        tag2 = ET.SubElement(root2, "Asset")
        tag3CH = ET.SubElement(tag2, "Metadata")
        ET.SubElement(
            tag3CH,
            "AMS",
            Version_Minor="0",
            Version_Major="1",
            Provider_ID=webProveedor,
            Provider=nameProveedor,
            Product="MOD",
            Description=title + " title asset",
            Creation_Date=hoy,
            Asset_Name=title + " title",
            Asset_ID=ID + str(i + 1),
            Asset_Class="title",
        )
        ET.SubElement(
            tag3CH,
            "App_Data",
            Value=actor1 + ", " + actor2 + ", " + actor3,
            Name="Actors_Display",
            App="MOD",
        )
        ET.SubElement(tag3CH, "App_Data", Value=actor1, Name="Actors", App="MOD")
        ET.SubElement(tag3CH, "App_Data", Value=actor2, Name="Actors", App="MOD")
        ET.SubElement(tag3CH, "App_Data", Value=actor3, Name="Actors", App="MOD")
        ET.SubElement(tag3CH, "App_Data", Value=durac, Name="Run_Time", App="MOD")
        ET.SubElement(
            tag3CH, "App_Data", Value=self.getClasif(clasif), Name="Rating", App="MOD"
        )
        ET.SubElement(
            tag3CH, "App_Data", Value=self.getGenero(genero), Name="Genre", App="MOD"
        )
        ET.SubElement(
            tag3CH, "App_Data", Value=i2VOD, Name="Licensing_Window_Start", App="MOD"
        )
        ET.SubElement(
            tag3CH, "App_Data", Value=f2VOD, Name="Licensing_Window_End", App="MOD"
        )
        ET.SubElement(
            tag3CH, "App_Data", Value="Ilapa", Name="Available_in_Localities", App="MOD"
        )
        ET.SubElement(tag3CH, "App_Data", Value=title, Name="Title_Brief", App="MOD")
        ET.SubElement(tag3CH, "App_Data", Value=title, Name="Title", App="MOD")
        ET.SubElement(
            tag3CH, "App_Data", Value=title, Name="Title_Sort_Name", App="MOD"
        )
        ET.SubElement(
            tag3CH, "App_Data", Value=sinopsis[:128], Name="Summary_Short", App="MOD"
        )
        ET.SubElement(
            tag3CH, "App_Data", Value=sinopsis, Name="Summary_Medium", App="MOD"
        )
        ET.SubElement(
            tag3CH, "App_Data", Value=sinopsis, Name="Summary_Long", App="MOD"
        )
        ET.SubElement(tag3CH, "App_Data", Value=str(year), Name="Year", App="MOD")
        ET.SubElement(tag3CH, "App_Data", Value=adult, Name="Adult", App="MOD")
        ET.SubElement(tag3CH, "App_Data", Value="", Name="Audience", App="MOD")
        ET.SubElement(tag3CH, "App_Data", Value=director, Name="Director", App="MOD")
        ET.SubElement(
            tag3CH,
            "App_Data",
            Value=duracionDisplay,
            Name="Display_Run_Time",
            App="MOD",
        )
        ET.SubElement(tag3CH, "App_Data", Value="", Name="Studio", App="MOD")
        ET.SubElement(
            tag3CH,
            "App_Data",
            Value=self.getCountry(country),
            Name="Country_of_Origin",
            App="MOD",
        )
        ET.SubElement(tag3CH, "App_Data", Value="", Name="Preview_Period", App="MOD")
        ET.SubElement(tag3CH, "App_Data", Value="", Name="Contract_Name", App="MOD")
        ET.SubElement(tag3CH, "App_Data", Value="title", Name="Type", App="MOD")
        ET.SubElement(tag3CH, "App_Data", Value="movie", Name="Show_Type", App="MOD")
        tag4CH = ET.SubElement(tag2, "Asset")
        tag4Children = ET.SubElement(tag4CH, "Metadata")
        ET.SubElement(
            tag4Children,
            "AMS",
            Version_Minor="0",
            Version_Major="1",
            Provider_ID=webProveedor,
            Provider=nameProveedor,
            Product="MOD",
            Description=title + " movie asset",
            Creation_Date=hoy,
            Asset_Name=title + " movie",
            Asset_ID=ID + str(i + 2),
            Asset_Class="movie",
        )
        ET.SubElement(tag4Children, "App_Data", Value="Movie", Name="Type", App="MOD")
        ET.SubElement(
            tag4Children, "App_Data", Value="50190", Name="Bit_Rate", App="MOD"
        )
        ET.SubElement(
            tag4Children, "App_Data", Value="Stereo", Name="Audio_Type", App="MOD"
        )
        ET.SubElement(
            tag4Children, "App_Data", Value="MPEG2", Name="Encoding_Type", App="MOD"
        )
        ET.SubElement(
            tag4Children,
            "App_Data",
            Value=str(self.getImgInfo(houseNumber + ".ts", getImg)[1]),
            Name="Content_CheckSum",
            App="MOD",
        )
        ET.SubElement(
            tag4Children,
            "App_Data",
            Value="Widescreen",
            Name="Screen_Format",
            App="MOD",
        )
        ET.SubElement(tag4Children, "App_Data", Value="es", Name="Languages", App="MOD")
        ET.SubElement(
            tag4Children,
            "App_Data",
            Value=str(self.getImgInfo(houseNumber + ".ts", getImg)[0]),
            Name="Content_FileSize",
            App="MOD",
        )
        ET.SubElement(
            tag4Children, "App_Data", Value=formato, Name="HDContent", App="MOD"
        )
        ET.SubElement(tag4CH, "Content", Value=houseNumber + ".ts")
        tag5CH = ET.SubElement(tag2, "Asset")
        tag5Children = ET.SubElement(tag5CH, "Metadata")
        ET.SubElement(
            tag5Children,
            "AMS",
            Version_Minor="0",
            Version_Major="1",
            Provider_ID=webProveedor,
            Provider=nameProveedor,
            Product="MOD",
            Description=title + " poster asset",
            Creation_Date=hoy,
            Asset_Name=title + " poster",
            Asset_ID=ID + str(i + 3),
            Asset_Class="poster",
        )
        ET.SubElement(
            tag5Children,
            "App_Data",
            Value="640x960",
            Name="Image_Aspect_Ratio",
            App="MOD",
        )
        ET.SubElement(
            tag5Children,
            "App_Data",
            Value=str(self.getImgInfo(houseNumber + "_1.jpg", getImg)[0]),
            Name="Content_FileSize",
            App="MOD",
        )
        ET.SubElement(
            tag5Children,
            "App_Data",
            Value=str(self.getImgInfo(houseNumber + "_1.jpg", getImg)[1]),
            Name="Content_CheckSum",
            App="MOD",
        )
        ET.SubElement(tag5Children, "App_Data", Value="poster", Name="Type", App="MOD")
        ET.SubElement(tag5CH, "Content", Value=houseNumber + "_1.jpg")

        # PRETTY xml ADI
        pretty = minidom.parseString(ET.tostring(root2)).toprettyxml(indent="   ")
        rot = ET.fromstring(pretty)
        doc2 = ET.ElementTree(rot)

        # wb string y bytes
        with open(path3 + "/" + houseNumber + ".xml", "wb") as f:
            f.write(
                '<?xml version="1.0" encoding="UTF-8" ?><!DOCTYPE ADI SYSTEM "ADI.DTD">'.encode(
                    "utf8"
                )
            )
            doc2.write(f, "utf-8")

    def megacable_meta(
        self,
        ID,
        nameProveedor,
        path1,
        houseNumber,
        country,
        title,
        sinopsis,
        actor1,
        actor2,
        actor3,
        director,
        genero,
        year,
        clasif,
        durac,
        duracionDisplay,
        formato,
        iVOD,
        fVOD,
    ):

        i = 30
        root1 = ET.Element("ADI")
        tag1 = ET.SubElement(root1, "Metadata")
        ET.SubElement(
            tag1,
            "AMS",
            Verb="",
            Asset_ID=ID[4:] + str(i),
            Asset_Class=" package",
            Provider_ID=nameProveedor,
        )

        tag2 = ET.SubElement(root1, "Asset")
        tag3CH = ET.SubElement(tag2, "Metadata")
        ET.SubElement(
            tag3CH,
            "AMS",
            Verb="",
            Asset_ID=ID[4:] + str(i + 1),
            Asset_Class=" title",
            Provider_ID=nameProveedor,
        )
        ET.SubElement(tag3CH, "App_Data ", Value="title", Name="Type", App="MOD")
        ET.SubElement(
            tag3CH, "App_Data ", Value=title, Name="title ", App="MOD", Language="es"
        )
        ET.SubElement(
            tag3CH, "App_Data ", Value=title, Name="title ", App="MOD", Language="en"
        )
        ET.SubElement(tag3CH, "App_Data ", Value=str(year), Name="year ", App="MOD")
        ET.SubElement(
            tag3CH,
            "App_Data ",
            Value=sinopsis[:1024],
            Name="Summary_Medium ",
            App="MOD",
            Language="es",
        )
        ET.SubElement(
            tag3CH,
            "App_Data ",
            Value=sinopsis[:1024],
            Name="Summary_Medium ",
            App="MOD",
            Language="en",
        )
        ET.SubElement(
            tag3CH, "App_Data ", Value=iVOD, Name="Licensing_Window_Start ", App="MOD"
        )
        ET.SubElement(
            tag3CH, "App_Data ", Value=fVOD, Name="Licensing_Window_End", App="MOD"
        )
        ET.SubElement(tag3CH, "App_Data ", Value="-1", Name="DestAreas ", App="")
        ET.SubElement(tag3CH, "App_Data ", Value=durac, Name="Run_Time ", App="MOD")
        ET.SubElement(
            tag3CH,
            "App_Data ",
            Value=duracionDisplay,
            Name="Display_Run_Time",
            App="MOD",
        )
        ET.SubElement(
            tag3CH,
            "App_Data ",
            Value=actor1 + ", " + actor2 + ", " + actor3,
            Name="Actors ",
            App="MOD",
            Language="es",
        )
        ET.SubElement(
            tag3CH,
            "App_Data ",
            Value=director,
            Name="Director ",
            App="MOD",
            Language="es",
        )
        ET.SubElement(
            tag3CH,
            "App_Data ",
            Value="",
            Name="Distributor_Name ",
            App="MOD",
            Language="es",
        )
        ET.SubElement(
            tag3CH,
            "App_Data ",
            Value="",
            Name="Distributor_Name ",
            App="MOD",
            Language="en",
        )
        ET.SubElement(
            tag3CH,
            "App_Data ",
            Value="",
            Name="Distributor_Type ",
            App="MOD",
            Language="es",
        )
        ET.SubElement(
            tag3CH,
            "App_Data ",
            Value="",
            Name="Distributor_Type ",
            App="MOD",
            Language="en",
        )
        ET.SubElement(
            tag3CH, "App_Data ", Value="", Name="Writers ", App="MOD", Language="es"
        )
        ET.SubElement(
            tag3CH,
            "App_Data ",
            Value=self.getCountry(country),
            Name="Country_of_Origin ",
            App="MOD",
        )
        ET.SubElement(
            tag3CH, "App_Data ", Value=self.getClasif(clasif), Name="Rating ", App="MOD"
        )
        ET.SubElement(
            tag3CH,
            "App_Data ",
            Value="MX",
            Name="DistributionList",
            App="MOD",
            Language="es",
        )
        ET.SubElement(
            tag3CH,
            "App_Data ",
            Name="DistributionList",
            App="MOD",
            Language="en",
            Valor="MX",
        )
        ET.SubElement(tag3CH, "App_Data ", Value="es", Name="Languages ", App="MOD")
        ET.SubElement(
            tag3CH, "App_Data ", Value="", Name="Subtitle_Languages ", App="MOD"
        )
        ET.SubElement(
            tag3CH, "App_Data ", Value=self.getGenero(genero), Name="Genre ", App="MOD"
        )
        ET.SubElement(
            tag3CH,
            "App_Data ",
            Value="Seachange On demand/OTT/En Renta/" + self.getGenero(genero),
            Name="Category ",
            App="MOD",
        )
        ET.SubElement(
            tag3CH,
            "App_Data ",
            Value="Seachange On demand/En Renta/" + self.getGenero(genero),
            Name="Category ",
            App="MOD",
        )
        ET.SubElement(
            tag3CH, "App_Data ", Value="0.00", Name="Suggested_Price ", App="MOD"
        )
        ET.SubElement(tag3CH, "App_Data ", Value="0", Name="Preview_Period ", App="MOD")
        ET.SubElement(tag3CH, "App_Data ", Value="48", Name="Rent_Period ", App="MOD")
        tag4CH = ET.SubElement(tag2, "Asset")
        tag4Children = ET.SubElement(tag4CH, "Metadata")
        ET.SubElement(
            tag4Children,
            "AMS",
            Verb="CREATE",
            Asset_ID=ID[4:] + str(i + 2),
            Asset_Class="movie",
            Provider=nameProveedor,
        )
        ET.SubElement(tag4Children, "App_Data ", Value="movie", Name="Type ", App="MOD")
        ET.SubElement(
            tag4Children, "App_Data ", Value="Y", Name="Encryption ", App="MOD"
        )
        ET.SubElement(
            tag4Children, "App_Data ", Value="1", Name="EncryptType ", App="MOD"
        )
        ET.SubElement(
            tag4Children, "App_Data ", Value="IPTV", Name="Domain ", App="MOD"
        )
        ET.SubElement(
            tag4Children, "App_Data ", Value="8000", Name="Bit_Rate ", App="MOD"
        )
        ET.SubElement(
            tag4Children, "App_Data ", Value="Stereo", Name="Audio_Type ", App="MOD"
        )
        ET.SubElement(
            tag4Children, "App_Data ", Value="N", Name="Copy_Protection ", App="MOD"
        )
        ET.SubElement(
            tag4Children,
            "App_Data ",
            Value="1",
            Name="Analog_Protection_System ",
            App="MOD",
        )
        ET.SubElement(
            tag4Children, "App_Data ", Value="Y", Name="Support_HDCP ", App="MOD"
        )
        ET.SubElement(tag4Children, "App_Data ", Value="3", Name="CGMS_A ", App="MOD")
        ET.SubElement(tag4CH, "Content", Value=houseNumber + ".ts")
        tag5CH = ET.SubElement(tag2, "Asset")
        tag5Children = ET.SubElement(tag5CH, "Metadata")
        ET.SubElement(
            tag5Children,
            "AMS",
            Verb="",
            Asset_ID=ID[4:] + str(i + 3),
            Asset_Class="movie",
            Provider=nameProveedor,
        )
        ET.SubElement(tag5Children, "App_Data ", Value="movie", Name="Type ", App="MOD")
        ET.SubElement(
            tag5Children, "App_Data ", Value="Y", Name="Encryption ", App="MOD"
        )
        ET.SubElement(
            tag5Children, "App_Data ", Value="2", Name="EncryptType ", App="MOD"
        )
        ET.SubElement(
            tag5Children, "App_Data ", Value="WEBTV", Name="Domain ", App="MOD"
        )
        ET.SubElement(
            tag5Children, "App_Data ", Value="8000", Name="Bit_Rate ", App="MOD"
        )
        ET.SubElement(
            tag5Children,
            "App_Data ",
            Value="3b31c6ca-c85d-4952-b54f-901289900ad3",
            Name="ProfileID ",
            App="MOD",
        )
        ET.SubElement(
            tag5Children, "App_Data ", Value="1", Name="Encoder_Mode ", App="MOD"
        )
        ET.SubElement(
            tag5Children, "App_Data ", Value="Stereo", Name="Audio_Type ", App="MOD"
        )
        ET.SubElement(
            tag5Children, "App_Data ", Value="N", Name="Copy_Protection ", App="MOD"
        )
        ET.SubElement(
            tag5Children,
            "App_Data ",
            Value="1",
            Name="Analog_Protection_System ",
            App="MOD",
        )
        ET.SubElement(
            tag5Children, "App_Data ", Value="Y", Name="Support_HDCP ", App="MOD"
        )
        ET.SubElement(tag5Children, "App_Data ", Value="3", Name="CGMS_A ", App="MOD")
        ET.SubElement(tag5CH, "Content", Value=houseNumber + ".ts")
        tag6CH = ET.SubElement(tag2, "Asset")
        tag6Children = ET.SubElement(tag6CH, "Metadata")
        ET.SubElement(
            tag6Children,
            "AMS",
            Verb="",
            Asset_ID=ID[4:] + str(i + 4),
            Asset_Class="poster",
            Provider=nameProveedor,
        )
        ET.SubElement(
            tag6Children, "App_Data ", Value="poster", Name="Type ", App="MOD"
        )
        ET.SubElement(tag6Children, "App_Data ", Value="1", Name="PType", App="MOD")
        ET.SubElement(tag6CH, "Content", Value=houseNumber + "_poster.jpg")
        tag7CH = ET.SubElement(tag2, "Asset")
        tag7Children = ET.SubElement(tag7CH, "Metadata")
        ET.SubElement(
            tag7Children,
            "AMS",
            Verb="",
            Asset_ID=ID[4:] + str(i + 5),
            Asset_Class="poster",
            Provider=nameProveedor,
        )
        ET.SubElement(
            tag7Children, "App_Data ", Value="poster", Name="Type ", App="MOD"
        )
        ET.SubElement(tag7Children, "App_Data ", Value="2", Name="PType", App="MOD")
        ET.SubElement(tag7CH, "Content", Value=houseNumber + "_wallpaper.jpg")
        tag8CH = ET.SubElement(tag2, "Asset")
        tag8Children = ET.SubElement(tag8CH, "Metadata")
        ET.SubElement(
            tag8Children,
            "AMS",
            Verb="",
            Asset_ID=ID[4:] + str(i + 6),
            Asset_Class="poster",
            Provider=nameProveedor,
        )
        ET.SubElement(
            tag8Children, "App_Data ", Value="poster", Name="Type ", App="MOD"
        )
        ET.SubElement(tag8Children, "App_Data ", Value="3", Name="PType", App="MOD")
        ET.SubElement(tag8CH, "Content", Value=houseNumber + "_poster.jpg")

        # pretty XML
        pretty = minidom.parseString(ET.tostring(root1)).toprettyxml(indent="   ")
        rot = ET.fromstring(pretty)
        doc1 = ET.ElementTree(rot)
        doc1.write(
            path1 + "/" + houseNumber + ".xml", encoding="utf-8", xml_declaration=True
        )

    def dish_meta(
        self,
        ID,
        getImg,
        callProveedor,
        nameProveedor,
        webProveedor,
        path2,
        houseNumber,
        title,
        sinopsis,
        actor1,
        actor2,
        actor3,
        director,
        genero,
        year,
        clasif,
        durac,
        duracionDisplay,
        formato,
        iVOD,
        fVOD,
    ):

        i = 20
        root = ET.Element("ADI")
        tag1 = ET.SubElement(root, "Metadata")
        ET.SubElement(
            tag1,
            "AMS",
            Asset_Class="package",
            Asset_ID=ID + str(i),
            Asset_Name=title + " package",
            Creation_Date="",
            Description=title + " package asset",
            Product="MOD",
            Provider_ID=webProveedor,
            Version_Major="1",
            Version_Minor="0",
        )
        ET.SubElement(
            tag1,
            "App_Data",
            App="MOD",
            Name="Metadata_Spec_Version",
            Value="CableLabsVOD1.1",
        )
        ET.SubElement(tag1, "App_Data", Name="Provider_Content_Tier", Value="CL1")
        tag2 = ET.SubElement(root, "Asset")
        tag3CH = ET.SubElement(tag2, "Metadata")
        ET.SubElement(
            tag3CH,
            "AMS",
            Asset_Class="title",
            Asset_ID=ID + str(i + 1),
            Asset_Name=title + " title",
            Creation_Date="",
            Description=title + " title asset",
            Product="MOD",
            Provider_ID=webProveedor,
            Version_Major="1",
            Version_Minor="0",
        )
        ET.SubElement(tag3CH, "App_Data", App="MOD", Name="Title_Brief", Value=title)
        ET.SubElement(
            tag3CH, "App_Data", App="MOD", Name="Closed_Captioning", Value="Y"
        )
        ET.SubElement(tag3CH, "App_Data", App="MOD", Name="Title", Value=title)
        ET.SubElement(tag3CH, "App_Data", App="MOD", Name="Type", Value="title")
        ET.SubElement(
            tag3CH, "App_Data", App="MOD", Name="Summary_Short", Value=sinopsis[:128]
        )
        # sinopsis[:sinopsis.rfind(",")]
        ET.SubElement(
            tag3CH, "App_Data", App="MOD", Name="Rating", Value=self.getClasif(clasif)
        )
        ET.SubElement(tag3CH, "App_Data", App="MOD", Name="Run_Time", Value=durac)
        ET.SubElement(
            tag3CH,
            "App_Data",
            App="MOD",
            Name="Display_Run_Time",
            Value=duracionDisplay,
        )
        ET.SubElement(tag3CH, "App_Data", App="MOD", Name="Year", Value=str(year))
        ET.SubElement(
            tag3CH,
            "App_Data",
            App="MOD",
            Name="Actors_Display",
            Value=actor1 + ", " + actor2 + ", " + actor3,
        )
        ET.SubElement(tag3CH, "App_Data", App="MOD", Name="Studio", Value=" ")
        ET.SubElement(tag3CH, "App_Data", App="MOD", Name="Director", Value=director)
        ET.SubElement(
            tag3CH, "App_Data", App="MOD", Name="Genre", Value=self.getGenero(genero)
        )
        ET.SubElement(tag3CH, "App_Data", App="MOD", Name="Billing_ID", Value="00000")
        ET.SubElement(
            tag3CH, "App_Data", App="MOD", Name="Maximum_Viewing_Length", Value=""
        )
        ET.SubElement(
            tag3CH, "App_Data", App="MOD", Name="Licensing_Window_Start", Value=iVOD
        )
        ET.SubElement(
            tag3CH, "App_Data", App="MOD", Name="Licensing_Window_End", Value=fVOD
        )
        ET.SubElement(
            tag3CH,
            "App_Data",
            App="MOD",
            Name="Provider_QA_Contact",
            Value="blanca.munoz@mvs.com",
        )
        ET.SubElement(
            tag3CH, "App_Data", App="MOD", Name="Suggested_Price", Value="0.00"
        )
        ET.SubElement(
            tag3CH, "App_Data", App="MOD", Name="Studio_Royalty_Percent", Value="0"
        )
        ET.SubElement(
            tag3CH, "App_Data", App="MOD", Name="Studio_Royalty_Minimum", Value="0"
        )
        ET.SubElement(
            tag3CH, "App_Data", App="MOD", Name="Studio_Royalty_Flat_Rate", Value="0"
        )
        ET.SubElement(tag3CH, "App_Data", App="MOD", Name="Studio_Name", Value=" ")
        ET.SubElement(tag3CH, "App_Data", App="MOD", Name="Preview_Period", Value="")
        ET.SubElement(
            tag3CH, "App_Data", App="MOD", Name="Distributor_Name", Value=nameProveedor
        )
        ET.SubElement(
            tag3CH,
            "App_Data",
            App="MOD",
            Name="Programmer_Call_Letters",
            Value=callProveedor,
        )
        tag4CH = ET.SubElement(tag2, "Asset")
        tag4Children = ET.SubElement(tag4CH, "Metadata")
        ET.SubElement(
            tag4Children,
            "AMS",
            Asset_Class="movie",
            Asset_ID=ID + str(i + 2),
            Asset_Name=title + " movie",
            Creation_Date="",
            Description=title + " movie asset",
            Product="MOD",
            Provider=nameProveedor,
            Provider_ID=webProveedor,
            Version_Major="1",
            Version_Minor="0",
        )
        ET.SubElement(tag4Children, "App_Data", App="MOD", Name="Type", Value="movie")
        ET.SubElement(tag4Children, "App_Data", App="MOD", Name="Languages", Value="es")
        ET.SubElement(
            tag4Children, "App_Data", App="MOD", Name="Audio_Type", Value="Stereo"
        )
        ET.SubElement(
            tag4Children,
            "App_Data",
            App="MOD",
            Name="Content_FileSize",
            Value=str(self.getImgInfo(houseNumber + ".mp4", getImg)[0]),
        )
        ET.SubElement(
            tag4Children,
            "App_Data",
            App="MOD",
            Name="Content_CheckSum",
            Value=str(self.getImgInfo(houseNumber + ".mp4", getImg)[1]),
        )
        ET.SubElement(
            tag4Children, "App_Data", App="MOD", Name="HDContent", Value=formato
        )
        ET.SubElement(
            tag4Children, "App_Data", App="MOD", Name="Bit_Rate", Value="50190"
        )
        ET.SubElement(tag4CH, "Content", Value=houseNumber + ".mpg")
        tag5CH = ET.SubElement(tag2, "Asset")
        tag5Children = ET.SubElement(tag5CH, "Metadata")
        ET.SubElement(
            tag5Children,
            "AMS",
            Asset_Class="poster",
            Asset_ID=ID + str(i + 3),
            Asset_Name=title + " movie",
            Creation_Date="",
            Description=title + " poster asset",
            Product="MOD",
            Provider=nameProveedor,
            Provider_ID=webProveedor,
            Version_Major="1",
            Version_Minor="0",
        )
        ET.SubElement(tag5Children, "App_Data", App="MOD", Name="Type", Value="poster")
        ET.SubElement(
            tag5Children,
            "App_Data",
            App="MOD",
            Name="Content_FileSize",
            Value=str(self.getImgInfo(houseNumber + "_1.jpg", getImg)[0]),
        )  # hash AQUIII
        ET.SubElement(
            tag5Children,
            "App_Data",
            App="MOD",
            Name="Content_CheckSum",
            Value=str(self.getImgInfo(houseNumber + "_1.jpg", getImg)[1]),
        )
        ET.SubElement(tag5CH, "Content", Value=str(houseNumber) + "_1.jpg")

        # Pretty XML
        pretty = minidom.parseString(ET.tostring(root)).toprettyxml(indent="   ")
        rot = ET.fromstring(pretty)
        doc = ET.ElementTree(rot)

        doc.write(
            path2 + "/" + houseNumber + ".xml", encoding="utf-8", xml_declaration=True
        )

    def cableonda_meta(
        self,
        ID,
        getImg,
        callProveedor,
        nameProveedor,
        webProveedor,
        path3,
        houseNumber,
        title,
        sinopsis,
        actor1,
        actor2,
        actor3,
        director,
        genero,
        year,
        clasif,
        durac,
        duracionDisplay,
        formato,
        i2VOD,
        f2VOD,
    ):

        today = date.today()
        hoy = today.strftime("%d/%m/%Y")

        i = 10

        root2 = ET.Element("ADI")
        tag1 = ET.SubElement(root2, "Metadata")
        ET.SubElement(
            tag1,
            "AMS",
            Version_Minor="0",
            Version_Major="1",
            Provider_ID=webProveedor,
            Provider=nameProveedor,
            Product="MOD",
            Description=title + " package asset",
            Creation_Date=hoy,
            Asset_Name=title + " package",
            Asset_ID=ID + str(i),
            Asset_Class="package",
        )
        ET.SubElement(
            tag1,
            "App_Data",
            Value="CableLabsVOD1.1",
            Name="Metadata_Spec_Version",
            App="MOD",
        )
        ET.SubElement(
            tag1,
            "App_Data",
            Value=callProveedor + "1",
            Name="Provider_Content_Tier",
            App="MOD",
        )
        tag2 = ET.SubElement(root2, "Asset")
        tag3CH = ET.SubElement(tag2, "Metadata")
        ET.SubElement(
            tag3CH,
            "AMS",
            Version_Minor="0",
            Version_Major="1",
            Provider_ID=webProveedor,
            Provider=nameProveedor,
            Product="MOD",
            Description=title + " title asset",
            Creation_Date=hoy,
            Asset_Name=title + " title",
            Asset_ID=ID + str(i + 1),
            Asset_Class="title",
        )
        ET.SubElement(tag3CH, "App_Data", Value=title, Name="Title_Brief", App="MOD")
        ET.SubElement(
            tag3CH, "App_Data", Value="Y", Name="Closed_Captioning", App="MOD"
        )
        ET.SubElement(tag3CH, "App_Data", Value=title, Name="Title", App="MOD")
        ET.SubElement(tag3CH, "App_Data", Value="title", Name="Type", App="MOD")
        ET.SubElement(
            tag3CH, "App_Data", Value=sinopsis[:128], Name="Summary_Short", App="MOD"
        )
        ET.SubElement(
            tag3CH, "App_Data", Value=self.getClasif(clasif), Name="Rating", App="MOD"
        )
        ET.SubElement(tag3CH, "App_Data", Value=durac, Name="Run_Time", App="MOD")
        ET.SubElement(
            tag3CH,
            "App_Data",
            Value=duracionDisplay,
            Name="Display_Run_Time",
            App="MOD",
        )
        ET.SubElement(tag3CH, "App_Data", Value=str(year), Name="Year", App="MOD")
        ET.SubElement(
            tag3CH,
            "App_Data",
            Value=actor1 + ", " + actor2 + ", " + actor3,
            Name="Actors_Display",
            App="MOD",
        )
        ET.SubElement(tag3CH, "App_Data", Value=" ", Name="Studio", App="MOD")
        ET.SubElement(tag3CH, "App_Data", Value=director, Name="Director", App="MOD")
        ET.SubElement(
            tag3CH, "App_Data", Value=self.getGenero(genero), Name="Genre", App="MOD"
        )
        ET.SubElement(tag3CH, "App_Data", Value="00000", Name="Billing_ID", App="MOD")
        ET.SubElement(
            tag3CH,
            "App_Data",
            Value="00:48:00",
            Name="Maximum_Viewing_Length",
            App="MOD",
        )
        ET.SubElement(
            tag3CH, "App_Data", Value=i2VOD, Name="Licensing_Window_Start", App="MOD"
        )
        ET.SubElement(
            tag3CH, "App_Data", Value=f2VOD, Name="Licensing_Window_End", App="MOD"
        )
        ET.SubElement(
            tag3CH,
            "App_Data",
            Value="blanca.munoz@mvs.com",
            Name="Provider_QA_Contact",
            App="MOD",
        )
        ET.SubElement(
            tag3CH, "App_Data", Value="0.00", Name="Suggested_Price", App="MOD"
        )
        ET.SubElement(
            tag3CH, "App_Data", Value="0", Name="Studio_Royalty_Percent", App="MOD"
        )
        ET.SubElement(
            tag3CH, "App_Data", Value="0", Name="Studio_Royalty_Minimum", App="MOD"
        )
        ET.SubElement(
            tag3CH, "App_Data", Value="0", Name="Studio_Royalty_Flat_Rate", App="MOD"
        )
        ET.SubElement(tag3CH, "App_Data", Value=" ", Name="Studio_Name", App="MOD")
        ET.SubElement(tag3CH, "App_Data", Value="180", Name="Preview_Period", App="MOD")
        ET.SubElement(
            tag3CH, "App_Data", Value=nameProveedor, Name="Distributor_Name", App="MOD"
        )
        ET.SubElement(
            tag3CH,
            "App_Data",
            Value=callProveedor,
            Name="Programmer_Call_Letters",
            App="MOD",
        )
        tag4CH = ET.SubElement(tag2, "Asset")
        tag4Children = ET.SubElement(tag4CH, "Metadata")
        ET.SubElement(
            tag4Children,
            "AMS",
            Version_Minor="0",
            Version_Major="1",
            Provider_ID=webProveedor,
            Provider=nameProveedor,
            Product="MOD",
            Description=title + " movie asset",
            Creation_Date=hoy,
            Asset_Name=title + " movie",
            Asset_ID=ID + str(i + 2),
            Asset_Class="movie",
        )
        ET.SubElement(tag4Children, "App_Data", Value="movie", Name="Type", App="MOD")
        ET.SubElement(tag4Children, "App_Data", Value="es", Name="Languages", App="MOD")
        ET.SubElement(
            tag4Children, "App_Data", Value="Stereo", Name="Audio_Type", App="MOD"
        )
        ET.SubElement(
            tag4Children,
            "App_Data",
            Value=str(self.getImgInfo(houseNumber + ".mp4", getImg)[0]),
            Name="Content_FileSize",
            App="MOD",
        )
        ET.SubElement(
            tag4Children,
            "App_Data",
            Value=str(self.getImgInfo(houseNumber + ".mp4", getImg)[1]),
            Name="Content_CheckSum",
            App="MOD",
        )
        ET.SubElement(
            tag4Children, "App_Data", Value=formato, Name="HDContent", App="MOD"
        )
        ET.SubElement(
            tag4Children, "App_Data", Value="50190", Name="Bit_Rate", App="MOD"
        )
        ET.SubElement(tag4CH, "Content", Value=houseNumber + ".mp4")
        tag5CH = ET.SubElement(tag2, "Asset")
        tag5Children = ET.SubElement(tag5CH, "Metadata")
        ET.SubElement(
            tag5Children,
            "AMS",
            Version_Minor="0",
            Version_Major="1",
            Provider_ID=webProveedor,
            Provider=nameProveedor,
            Product="MOD",
            Description=title + " poster asset",
            Creation_Date=hoy,
            Asset_Name=title + " poster",
            Asset_ID=ID + str(i + 3),
            Asset_Class="poster",
        )
        ET.SubElement(tag5Children, "App_Data", Value="poster", Name="Type", App="MOD")
        ET.SubElement(
            tag5Children,
            "App_Data",
            Value=str(self.getImgInfo(houseNumber + "_1.jpg", getImg)[0]),
            Name="Content_FileSize",
            App="MOD",
        )
        ET.SubElement(
            tag5Children,
            "App_Data",
            Value=str(self.getImgInfo(houseNumber + "_1.jpg", getImg)[1]),
            Name="Content_CheckSum",
            App="MOD",
        )
        ET.SubElement(tag5CH, "Content", Value=houseNumber + "_1.jpg")
        tag6CH = ET.SubElement(tag2, "Asset")
        tag6Children = ET.SubElement(tag6CH, "Metadata")
        ET.SubElement(
            tag6Children,
            "AMS",
            Version_Minor="0",
            Version_Major="1",
            Provider_ID=webProveedor,
            Provider=nameProveedor,
            Product="MOD",
            Description=title + " poster asset",
            Creation_Date=hoy,
            Asset_Name=title + " poster",
            Asset_ID=ID + str(i + 4),
            Asset_Class="poster",
        )
        ET.SubElement(tag6Children, "App_Data", Value="poster", Name="Type", App="MOD")
        ET.SubElement(
            tag6Children,
            "App_Data",
            Value=str(self.getImgInfo(houseNumber + "_2.jpg", getImg)[0]),
            Name="Content_FileSize",
            App="MOD",
        )
        ET.SubElement(
            tag6Children,
            "App_Data",
            Value=str(self.getImgInfo(houseNumber + "_2.jpg", getImg)[1]),
            Name="Content_CheckSum",
            App="MOD",
        )
        ET.SubElement(tag6CH, "Content", Value=houseNumber + "_2.jpg")

        # PRETTY xml ADI
        pretty = minidom.parseString(ET.tostring(root2)).toprettyxml(indent="   ")
        rot = ET.fromstring(pretty)
        doc2 = ET.ElementTree(rot)

        # wb string y bytes
        with open(path3 + "/" + houseNumber + ".xml", "wb") as f:
            f.write(
                '<?xml version="1.0" encoding="UTF-8" ?><!DOCTYPE ADI SYSTEM "ADI.DTD">'.encode(
                    "utf8"
                )
            )
            doc2.write(f, "utf-8")

    def youtube_meta(
        self,
        ID,
        getImg,
        callProveedor,
        nameProveedor,
        webProveedor,
        path3,
        houseNumber,
        title,
        sinopsis,
        actor1,
        actor2,
        actor3,
        director,
        genero,
        year,
        clasif,
        durac,
        duracionDisplay,
        formato,
        i2VOD,
        f2VOD,
        vM,
        cc,
    ):

        today = date.today()
        hoy = today.strftime("%Y-%m-%d")
        # hoy = today.strftime("%d/%m/%Y")

        i = 10

        root2 = ET.Element("ADI")
        tag1 = ET.SubElement(root2, "Metadata")
        ET.SubElement(
            tag1,
            "AMS",
            Version_Minor="0",
            Version_Major=vM,
            Provider_ID=webProveedor,
            Provider=nameProveedor,
            Product="MOD",
            Description=title + " package asset",
            Creation_Date=hoy,
            Asset_Name=title + " package",
            Asset_ID=ID + str(i),
            Asset_Class="package",
        )
        ET.SubElement(
            tag1,
            "App_Data",
            Value="CableLabsVOD1.1",
            Name="Metadata_Spec_Version",
            App="MOD",
        )
        ET.SubElement(
            tag1,
            "App_Data",
            Value=callProveedor + "1",
            Name="Provider_Content_Tier",
            App="MOD",
        )
        tag2 = ET.SubElement(root2, "Asset")
        tag3CH = ET.SubElement(tag2, "Metadata")
        ET.SubElement(
            tag3CH,
            "AMS",
            Version_Minor="0",
            Version_Major=vM,
            Provider_ID=webProveedor,
            Provider=nameProveedor,
            Product="MOD",
            Description=title + " title asset",
            Creation_Date=hoy,
            Asset_Name=title + " title",
            Asset_ID=ID + str(i + 1),
            Asset_Class="title",
        )
        ET.SubElement(tag3CH, "App_Data", Value=title, Name="Title_Brief", App="MOD")
        ET.SubElement(tag3CH, "App_Data", Value=cc, Name="Closed_Captioning", App="MOD")
        ET.SubElement(tag3CH, "App_Data", Value=title, Name="Title", App="MOD")
        ET.SubElement(tag3CH, "App_Data", Value="title", Name="Type", App="MOD")
        ET.SubElement(
            tag3CH, "App_Data", Value=sinopsis[:128], Name="Summary_Short", App="MOD"
        )
        ET.SubElement(
            tag3CH, "App_Data", Value=self.getClasif(clasif), Name="Rating", App="MOD"
        )
        ET.SubElement(tag3CH, "App_Data", Value=durac, Name="Run_Time", App="MOD")
        ET.SubElement(
            tag3CH,
            "App_Data",
            Value=duracionDisplay,
            Name="Display_Run_Time",
            App="MOD",
        )
        ET.SubElement(tag3CH, "App_Data", Value=str(year), Name="Year", App="MOD")
        ET.SubElement(
            tag3CH,
            "App_Data",
            Value=self.comilla(actor1)
            + ", "
            + self.comilla(actor2)
            + ", "
            + self.comilla(actor3),
            Name="Actors_Display",
            App="MOD",
        )
        ET.SubElement(tag3CH, "App_Data", Value=" ", Name="Studio", App="MOD")
        ET.SubElement(tag3CH, "App_Data", Value=director, Name="Director", App="MOD")
        ET.SubElement(tag3CH, "App_Data", Value="Movie", Name="Show_Type", App="MOD")
        ET.SubElement(
            tag3CH, "App_Data", Value=self.getGenero(genero), Name="Genre", App="MOD"
        )
        ET.SubElement(tag3CH, "App_Data", Value="00000", Name="Billing_ID", App="MOD")
        ET.SubElement(
            tag3CH,
            "App_Data",
            Value="00:48:00",
            Name="Maximum_Viewing_Length",
            App="MOD",
        )
        ET.SubElement(
            tag3CH, "App_Data", Value=i2VOD, Name="Licensing_Window_Start", App="MOD"
        )
        ET.SubElement(
            tag3CH, "App_Data", Value=f2VOD, Name="Licensing_Window_End", App="MOD"
        )
        ET.SubElement(
            tag3CH,
            "App_Data",
            Value="blanca.munoz@mvs.com",
            Name="Provider_QA_Contact",
            App="MOD",
        )
        ET.SubElement(
            tag3CH, "App_Data", Value="0.00", Name="Suggested_Price", App="MOD"
        )
        ET.SubElement(
            tag3CH, "App_Data", Value="0", Name="Studio_Royalty_Percent", App="MOD"
        )
        ET.SubElement(
            tag3CH, "App_Data", Value="0", Name="Studio_Royalty_Minimum", App="MOD"
        )
        ET.SubElement(
            tag3CH, "App_Data", Value="0", Name="Studio_Royalty_Flat_Rate", App="MOD"
        )
        ET.SubElement(tag3CH, "App_Data", Value=" ", Name="Studio_Name", App="MOD")
        ET.SubElement(tag3CH, "App_Data", Value="180", Name="Preview_Period", App="MOD")
        ET.SubElement(
            tag3CH, "App_Data", Value=nameProveedor, Name="Distributor_Name", App="MOD"
        )
        ET.SubElement(
            tag3CH,
            "App_Data",
            Value=callProveedor,
            Name="Programmer_Call_Letters",
            App="MOD",
        )
        tag4CH = ET.SubElement(tag2, "Asset")
        tag4Children = ET.SubElement(tag4CH, "Metadata")
        ET.SubElement(
            tag4Children,
            "AMS",
            Version_Minor="0",
            Version_Major=vM,
            Provider_ID=webProveedor,
            Provider=nameProveedor,
            Product="MOD",
            Description=title + " movie asset",
            Creation_Date=hoy,
            Asset_Name=title + " movie",
            Asset_ID=ID + str(i + 2),
            Asset_Class="movie",
        )
        ET.SubElement(tag4Children, "App_Data", Value="movie", Name="Type", App="MOD")
        ET.SubElement(tag4Children, "App_Data", Value="es", Name="Languages", App="MOD")
        ET.SubElement(
            tag4Children, "App_Data", Value="Stereo", Name="Audio_Type", App="MOD"
        )
        ET.SubElement(
            tag4Children,
            "App_Data",
            Value=str(self.getImgInfo(houseNumber + ".mp4", getImg)[0]),
            Name="Content_FileSize",
            App="MOD",
        )
        ET.SubElement(
            tag4Children,
            "App_Data",
            Value=str(self.getImgInfo(houseNumber + ".mp4", getImg)[1]),
            Name="Content_CheckSum",
            App="MOD",
        )
        ET.SubElement(
            tag4Children, "App_Data", Value=formato, Name="HDContent", App="MOD"
        )
        ET.SubElement(
            tag4Children, "App_Data", Value="50190", Name="Bit_Rate", App="MOD"
        )
        ET.SubElement(tag4CH, "Content", Value=houseNumber + ".mp4")
        tag5CH = ET.SubElement(tag2, "Asset")
        tag5Children = ET.SubElement(tag5CH, "Metadata")
        ET.SubElement(
            tag5Children,
            "AMS",
            Version_Minor="0",
            Version_Major=vM,
            Provider_ID=webProveedor,
            Provider=nameProveedor,
            Product="MOD",
            Description=title + " poster asset",
            Creation_Date=hoy,
            Asset_Name=title + " poster",
            Asset_ID=ID + str(i + 3),
            Asset_Class="poster",
        )
        ET.SubElement(tag5Children, "App_Data", Value="poster", Name="Type", App="MOD")
        ET.SubElement(
            tag5Children,
            "App_Data",
            Value=str(self.getImgInfo(houseNumber + "_1.jpg", getImg)[0]),
            Name="Content_FileSize",
            App="MOD",
        )
        ET.SubElement(
            tag5Children,
            "App_Data",
            Value=str(self.getImgInfo(houseNumber + "_1.jpg", getImg)[1]),
            Name="Content_CheckSum",
            App="MOD",
        )
        ET.SubElement(tag5CH, "Content", Value=houseNumber + "_1.jpg")
        tag6CH = ET.SubElement(tag2, "Asset")
        tag6Children = ET.SubElement(tag6CH, "Metadata")
        ET.SubElement(
            tag6Children,
            "AMS",
            Version_Minor="0",
            Version_Major=vM,
            Provider_ID=webProveedor,
            Provider=nameProveedor,
            Product="MOD",
            Description=title + " poster asset",
            Creation_Date=hoy,
            Asset_Name=title + " poster",
            Asset_ID=ID + str(i + 4),
            Asset_Class="poster",
        )
        ET.SubElement(tag6Children, "App_Data", Value="poster", Name="Type", App="MOD")
        ET.SubElement(
            tag6Children,
            "App_Data",
            Value=str(self.getImgInfo(houseNumber + "_2.jpg", getImg)[0]),
            Name="Content_FileSize",
            App="MOD",
        )
        ET.SubElement(
            tag6Children,
            "App_Data",
            Value=str(self.getImgInfo(houseNumber + "_2.jpg", getImg)[1]),
            Name="Content_CheckSum",
            App="MOD",
        )
        ET.SubElement(tag6CH, "Content", Value=houseNumber + "_2.jpg")
        tag7CH = ET.SubElement(tag2, "Asset")
        tag7Children = ET.SubElement(tag7CH, "Metadata")
        ET.SubElement(
            tag7Children,
            "AMS",
            Version_Minor="0",
            Version_Major=vM,
            Provider_ID=webProveedor,
            Provider=nameProveedor,
            Product="MOD",
            Description=title + " closed caption",
            Creation_Date=hoy,
            Asset_Name=title + " closed caption",
            Asset_ID=ID + str(i + 4),
            Asset_Class="closed caption",
        )
        ET.SubElement(
            tag7Children, "App_Data", Value="closed caption", Name="Type", App="MOD"
        )
        ET.SubElement(
            tag7Children,
            "App_Data",
            Value=str(self.getImgInfo(houseNumber + ".scc", getImg)[0]),
            Name="Content_FileSize",
            App="MOD",
        )
        ET.SubElement(
            tag7Children,
            "App_Data",
            Value=str(self.getImgInfo(houseNumber + ".scc", getImg)[1]),
            Name="Content_CheckSum",
            App="MOD",
        )
        ET.SubElement(tag7CH, "Content", Value=houseNumber + ".scc")

        # PRETTY xml ADI
        pretty = minidom.parseString(ET.tostring(root2)).toprettyxml(indent="   ")
        rot = ET.fromstring(pretty)
        doc2 = ET.ElementTree(rot)

        # wb string y bytes
        with open(path3 + "/" + houseNumber + ".xml", "wb") as f:
            f.write(
                '<?xml version="1.0" encoding="UTF-8" ?><!DOCTYPE ADI SYSTEM "ADI.DTD">'.encode(
                    "utf8"
                )
            )
            doc2.write(f, "utf-8")

    def tigo_meta(
        self,
        ID,
        getImg,
        callProveedor,
        country,
        nameProveedor,
        webProveedor,
        path4,
        houseNumber,
        title,
        sinopsis,
        actor1,
        actor2,
        actor3,
        director,
        genero,
        year,
        clasif,
        durac,
        duracionDisplay,
        formato,
        iVOD,
        fVOD,
    ):

        today = date.today()
        hoy = today.strftime("%d/%m/%Y")

        i = 40
        print(hoy)

        root2 = ET.Element("ADI")
        tag1 = ET.SubElement(root2, "Metadata")
        ET.SubElement(
            tag1,
            "AMS",
            Asset_Class="package",
            Asset_ID=ID + str(i),
            Asset_Name=title + " package",
            Creation_Date=hoy,
            Description=sinopsis[:128] + " package asset",
            Product="SVOD",
            Provider=nameProveedor,
            Provider_ID=webProveedor,
            Version_Minor="0",
            Version_Major="1",
        )
        ET.SubElement(
            tag1,
            "App_Data",
            App="SVOD",
            Name="Metadata_Spec_Version",
            Value="CableLabsVOD1.1",
        )
        tag2 = ET.SubElement(root2, "Asset")
        tag3CH = ET.SubElement(tag2, "Metadata")
        ET.SubElement(
            tag3CH,
            "AMS",
            Version_Minor="0",
            Version_Major="1",
            Provider_ID=webProveedor,
            Provider=nameProveedor,
            Product="SVOD",
            Description=sinopsis[:128] + " title asset",
            Creation_Date="",
            Asset_Name=title + " title",
            Asset_ID=ID + str(i + 1),
            Asset_Class="title",
        )
        ET.SubElement(tag3CH, "App_Data", Value="title", Name="Type", App="SVOD")
        ET.SubElement(tag3CH, "App_Data", Value=title, Name="Title", App="SVOD")
        ET.SubElement(
            tag3CH, "App_Data", Value=title, Name="Title_Sort_Name", App="SVOD"
        )
        ET.SubElement(tag3CH, "App_Data", Value=title, Name="Title_Brief", App="SVOD")
        ET.SubElement(
            tag3CH,
            "App_Data",
            Value=self.getClasifTigo(clasif),
            Name="Rating",
            App="SVOD",
        )
        ET.SubElement(
            tag3CH, "App_Data", Value=sinopsis[:128], Name="Summary_Short", App="SVOD"
        )
        ET.SubElement(
            tag3CH, "App_Data", Value=sinopsis[:1024], Name="Summary_Medium", App="SVOD"
        )
        ET.SubElement(
            tag3CH, "App_Data", Value=sinopsis, Name="Summary_Long", App="SVOD"
        )
        ET.SubElement(tag3CH, "App_Data", Value=durac, Name="Run_Time", App="SVOD")
        ET.SubElement(
            tag3CH,
            "App_Data",
            Value=duracionDisplay,
            Name="Display_Run_Time",
            App="SVOD",
        )
        ET.SubElement(tag3CH, "App_Data", Value=str(year), Name="Year", App="SVOD")
        ET.SubElement(tag3CH, "App_Data", Value=actor1, Name="Actors", App="SVOD")
        ET.SubElement(tag3CH, "App_Data", Value=actor2, Name="Actors", App="SVOD")
        ET.SubElement(tag3CH, "App_Data", Value=actor3, Name="Actors", App="SVOD")
        ET.SubElement(
            tag3CH,
            "App_Data",
            Value=actor1 + ", " + actor2 + ", " + actor3,
            Name="Actors_Display",
            App="SVOD",
        )
        ET.SubElement(
            tag3CH, "App_Data", Value=self.getGenero(genero), Name="Genre", App="SVOD"
        )
        ET.SubElement(
            tag3CH, "App_Data", Value="Gratis/Películas ", Name="Category", App="SVOD"
        )
        ET.SubElement(
            tag3CH, "App_Data", Value=iVOD, Name="Licensing_Window_Start", App="SVOD"
        )
        ET.SubElement(
            tag3CH, "App_Data", Value=fVOD, Name="Licensing_Window_End", App="SVOD"
        )
        ET.SubElement(
            tag3CH,
            "App_Data",
            Value="gt,sv,hn,cr,co,bo,py",
            Name="Available_in_countries",
            App="SVOD",
        )
        ET.SubElement(
            tag3CH, "App_Data", Value="SVOD_BASICO", Name="Package_offer_ID", App="SVOD"
        )
        ET.SubElement(tag3CH, "App_Data", Value="Movie", Name="Show_Type", App="SVOD")
        ET.SubElement(
            tag3CH, "App_Data", Value="1", Name="Propagation_Priority", App="SVOD"
        )
        ET.SubElement(tag3CH, "App_Data", Value="7", Name="Display_As_New", App="SVOD")
        ET.SubElement(
            tag3CH, "App_Data", Value="7", Name="Display_As_Last_Chance", App="SVOD"
        )
        ET.SubElement(
            tag3CH, "App_Data", Value="300", Name="Preview_Period", App="SVOD"
        )
        ET.SubElement(
            tag3CH,
            "App_Data ",
            Value=self.getCountry(country),
            Name="Country_of_Origin ",
            App="MOD",
        )
        ET.SubElement(
            tag3CH, "App_Data", Value="N", Name="Closed_Captioning", App="SVOD"
        )
        ET.SubElement(
            tag3CH,
            "App_Data",
            Value="blanca.munoz@mvs.com",
            Name="Provider_QA_Contact",
            App="SVOD",
        )
        ET.SubElement(
            tag3CH, "App_Data", Value="Studio Canal ", Name="Studio", App="SVOD"
        )
        ET.SubElement(
            tag3CH,
            "App_Data",
            Value="00:48:00",
            Name="Maximum_Viewing_Length",
            App="SVOD",
        )

        ET.SubElement(
            tag3CH, "App_Data", Value=nameProveedor, Name="Distributor_Name", App="SVOD"
        )
        ET.SubElement(
            tag3CH,
            "App_Data",
            Value=callProveedor,
            Name="Programmer_Call_Letters",
            App="SVOD",
        )

        tag4CH = ET.SubElement(tag2, "Asset")
        tag4Children = ET.SubElement(tag4CH, "Metadata")
        ET.SubElement(
            tag4Children,
            "AMS",
            Version_Minor="0",
            Version_Major="1",
            Provider_ID=webProveedor,
            Provider=nameProveedor,
            Product="SVOD",
            Description=sinopsis[:128],
            Creation_Date=hoy,
            Asset_Name=title,
            Asset_ID=ID + str(i + 2),
            Asset_Class="movie",
        )
        ET.SubElement(tag4Children, "App_Data", Value="movie", Name="Type", App="SVOD")
        ET.SubElement(
            tag4Children, "App_Data", Value="N", Name="Copy_Protection", App="SVOD"
        )
        ET.SubElement(
            tag4Children, "App_Data", Value="N", Name="Encryption", App="SVOD"
        )
        ET.SubElement(
            tag4Children, "App_Data", Value="Stereo", Name="Audio_Type", App="SVOD"
        )
        ET.SubElement(
            tag4Children, "App_Data", Value="es", Name="Languages", App="SVOD"
        )
        ET.SubElement(
            tag4Children, "App_Data", Value=formato, Name="HDContent", App="SVOD"
        )
        ET.SubElement(
            tag4Children,
            "App_Data",
            Value="Widescreen",
            Name="Screen_Format",
            App="SVOD",
        )
        ET.SubElement(
            tag4Children,
            "App_Data",
            Value=str(self.getImgInfo(houseNumber + ".mp4", getImg)[0]),
            Name="Content_FileSize",
            App="SVOD",
        )
        ET.SubElement(
            tag4Children,
            "App_Data",
            Value=str(self.getImgInfo(houseNumber + ".mp4", getImg)[1]),
            Name="Content_CheckSum",
            App="SVOD",
        )

        ET.SubElement(tag4CH, "Content", Value=houseNumber + ".mp4")
        tag5CH = ET.SubElement(tag2, "Asset")
        tag5Children = ET.SubElement(tag5CH, "Metadata")
        ET.SubElement(
            tag5Children,
            "AMS",
            Version_Minor="0",
            Version_Major="1",
            Provider_ID=webProveedor,
            Provider=nameProveedor,
            Product="SVOD",
            Description=sinopsis[:128] + " poster asset",
            Creation_Date=hoy,
            Asset_Name=title + " poster",
            Asset_ID=ID + str(i + 3),
            Asset_Class="poster",
        )
        ET.SubElement(tag5Children, "App_Data", Value="poster", Name="Type", App="SVOD")
        ET.SubElement(
            tag5Children,
            "App_Data",
            Value="640x480",
            Name="Image_Aspect_Ratio",
            App="SVOD",
        )
        ET.SubElement(
            tag5Children,
            "App_Data",
            Value=str(self.getImgInfo(houseNumber + "_1.jpg", getImg)[0]),
            Name="Content_FileSize",
            App="SVOD",
        )
        ET.SubElement(
            tag5Children,
            "App_Data",
            Value=str(self.getImgInfo(houseNumber + "_1.jpg", getImg)[1]),
            Name="Content_CheckSum",
            App="SVOD",
        )
        ET.SubElement(tag5CH, "Content", Value=houseNumber + "_1.jpg")

        # PRETTY xml ADI
        pretty = minidom.parseString(ET.tostring(root2)).toprettyxml(indent="   ")
        rot = ET.fromstring(pretty)
        doc2 = ET.ElementTree(rot)

        # wb string y bytes
        with open(path4 + "/" + ID + str(i) + ".xml", "wb") as f:
            f.write(
                '<?xml version="1.0" encoding="UTF-8" ?><!DOCTYPE ADI SYSTEM "ADI.DTD">'.encode(
                    "utf8"
                )
            )
            doc2.write(f, "utf-8")

    def getImgInfo(self, houseNumb, getImg):

        window.title("Trabajando en " + houseNumb)
        directorio = os.listdir(self.getImg)
        fileSz = ""
        hashData = ""
        for a in directorio:

            if not a.startswith("."):
                getSub = os.path.join(self.getImg, a)
                subdirectorio = os.listdir(getSub)

                for b in subdirectorio:

                    if not b.startswith("."):
                        if houseNumb in b:
                            Data = str(self.getImg + os.sep + a + os.sep + b)
                            hash_md5 = hashlib.md5()

                            with open(Data, "rb") as f:

                                print("generando MD5....")
                                for chunk in iter(lambda: f.read(16384), b""):
                                    hash_md5.update(chunk)

                            hashData = hash_md5.hexdigest()
                            # print(hashData)
                            fileSz = str(os.stat(Data).st_size)

        return fileSz, hashData

    def getGenero(self, genero):

        if genero == "Comedia":
            genero = "Comedy"

        if genero == "Comedia romantica":
            genero = "Romantic Comedy"

        if genero == "Adultos":
            genero = "Adult"

        if genero == "Suspenso":
            genero = "Suspense"

        if genero == "Acción":
            genero = "Action"

        if genero == "Infantil":
            genero = "Kids"

        if genero == "Misterio":
            genero = "Mystery"

        return genero

    def getClasif(self, clasif):

        if clasif == "A" or clasif == "AA":
            clasif = "G"

        if clasif == "B":
            clasif = "PG"

        if clasif == "B-15":
            clasif = "PG-13"

        if clasif == "C" or clasif == "D":
            clasif = "R"

        return clasif

    def getClasifTigo(self, clasif):
        print(clasif)
        if clasif == "A" or clasif == "AA":
            clasif = "general"
        elif clasif == "B":
            clasif = "+12"
        elif clasif == "B-15":
            clasif = "+18"
        elif clasif == "C" or clasif == "D":
            clasif = "adult"

        print(clasif)

        return clasif

    def getCountry(self, country):

        if country == "México":
            country = "MX"

        if country == "USA" or country == "Estados Unidos de América":
            country = "US"

        if country == "Perú":
            country = "PE"

        if country == "Colombia":
            country = "CO"

        if country == "Bolivia":
            country = "BO"

        if country == "Netherlands":
            country = "NL"

        return country

    def comilla(self, data):
        caracteres = '"'

        for x in range(len(caracteres)):
            data = data.replace(caracteres[x], "'")

        return data


def main():
    Metadata()


if __name__ == "__main__":
    main()
